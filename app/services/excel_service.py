"""Excel (XLSX) service para o app.

Gera planilhas que espelham a estrutura visual dos documentos (mesmas seções,
rótulos e dados do PDF). Reaproveita os mapeamentos de pdf_service para garantir
que nenhum campo seja perdido entre PDF e XLSX.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.pdf_service import _map_habitacao_to_template

# ---------------------------------------------------------------------------
# Estilos compartilhados (espelham o template ficha_habitacional.html)
# ---------------------------------------------------------------------------

_GRID = 8  # número de colunas do layout (A..H)

_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FILL_SECTION = PatternFill("solid", fgColor="A5A5A5")   # cabeçalho de seção (cinza)
_FILL_HEADER = PatternFill("solid", fgColor="D9D9D9")     # cabeçalho de tabela (cinza claro)

_FONT_BASE = Font(name="Calibri", size=10)
_FONT_BOLD = Font(name="Calibri", size=10, bold=True)
_FONT_SECTION = Font(name="Calibri", size=10, bold=True)

_AL_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_AL_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_range(ws, row: int, col_start: int, col_end: int, *, fill=None,
                 font=None, align=None):
    """Aplica borda (+ estilo opcional) a todas as células de um intervalo numa linha."""
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = _BORDER
        if fill is not None:
            cell.fill = fill
        if font is not None:
            cell.font = font
        if align is not None:
            cell.alignment = align


def _section(ws, row: int, title: str) -> int:
    """Cabeçalho de seção em barra cinza ocupando toda a largura. Retorna próxima linha."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_GRID)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = _FONT_SECTION
    cell.alignment = _AL_CENTER
    _style_range(ws, row, 1, _GRID, fill=_FILL_SECTION, font=_FONT_SECTION, align=_AL_CENTER)
    return row + 1


def _kv(ws, row: int, label: str, value, *, label_span: int = 2,
        value_span: int | None = None, fill_label=None) -> int:
    """Linha rótulo/valor. label ocupa label_span colunas, valor ocupa o restante.

    Retorna a próxima linha livre.
    """
    value_end = _GRID if value_span is None else label_span + value_span
    # rótulo
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=label_span)
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = _FONT_BOLD
    lc.alignment = _AL_LEFT
    _style_range(ws, row, 1, label_span, fill=fill_label, font=_FONT_BOLD, align=_AL_LEFT)
    # valor
    ws.merge_cells(start_row=row, start_column=label_span + 1, end_row=row, end_column=value_end)
    vc = ws.cell(row=row, column=label_span + 1, value=_clean(value))
    vc.font = _FONT_BASE
    vc.alignment = _AL_LEFT
    _style_range(ws, row, label_span + 1, value_end, font=_FONT_BASE, align=_AL_LEFT)
    return row + 1


def _clean(value):
    """Normaliza valor para célula: None/'' viram string vazia."""
    if value is None:
        return ""
    return value


def _row_cells(ws, row: int, values, *, fill=None, font=None, align=None,
               starts=None) -> int:
    """Escreve uma sequência de células lado a lado (1 col cada por padrão).

    `starts` opcional: lista de (col_start, col_end) para merges por célula.
    """
    font = font or _FONT_BASE
    align = align or _AL_CENTER
    if starts is None:
        for i, v in enumerate(values):
            cell = ws.cell(row=row, column=i + 1, value=_clean(v))
            cell.font = font
            cell.alignment = align
            cell.border = _BORDER
            if fill is not None:
                cell.fill = fill
    else:
        for (cs, ce), v in zip(starts, values):
            ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
            cell = ws.cell(row=row, column=cs, value=_clean(v))
            cell.font = font
            cell.alignment = align
            _style_range(ws, row, cs, ce, fill=fill, font=font, align=align)
    return row + 1


# ---------------------------------------------------------------------------
# Ficha habitacional
# ---------------------------------------------------------------------------

def gerar_xlsx_habitacao(ficha: dict) -> bytes:
    """Gera XLSX espelhando a ficha habitacional (mesmas seções do PDF)."""
    d = _map_habitacao_to_template(ficha)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ficha Habitacional"
    ws.sheet_view.showGridLines = False

    # larguras de coluna (8 colunas)
    widths = [20, 16, 16, 14, 14, 16, 12, 12]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    r = 1

    # Título / cabeçalho de entrevista
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=_GRID)
    title = ws.cell(row=r, column=1, value="FICHA HABITACIONAL — PMCMV")
    title.font = Font(name="Calibri", size=13, bold=True)
    title.alignment = _AL_CENTER
    r += 1
    if d.get("empreendimento"):
        r = _kv(ws, r, "EMPREENDIMENTO:", d.get("empreendimento"))
    r = _row_cells(
        ws, r,
        ["ENTREVISTA: ____/____/________", "ENTREVISTADOR: ______________________________"],
        font=_FONT_BOLD, align=_AL_LEFT,
        starts=[(1, 3), (4, _GRID)],
    )
    r += 1  # linha em branco

    # --- DADOS PARTICIPANTES ---
    r = _section(ws, r, "DADOS PARTICIPANTES")
    r = _row_cells(
        ws, r,
        ["1.PROPONENTE/IDADE", d.get("proponente1_nome"),
         (f"{d['proponente1_idade']} ANOS" if d.get("proponente1_idade") else ""),
         "CPF:", d.get("proponente1_cpf")],
        align=_AL_LEFT,
        starts=[(1, 2), (3, 5), (6, 6), (7, 7), (8, 8)],
    )
    r = _row_cells(
        ws, r,
        ["2.COOBRIGADO/IDADE", d.get("coobrigado_nome"), "",
         "CPF:", d.get("coobrigado_cpf")],
        align=_AL_LEFT,
        starts=[(1, 2), (3, 5), (6, 6), (7, 7), (8, 8)],
    )
    r = _kv(ws, r, "DEPENDENTES:", d.get("dependentes"))
    r = _kv(ws, r, "CONTATOS (TELEFONE):", d.get("contato_telefone"))
    r = _kv(ws, r, "CONTATOS (E-MAIL):", d.get("contato_email"))
    r += 1

    # --- RENDIMENTO E VINCULO EMPREGATICIOS ---
    r = _section(ws, r, "RENDIMENTO E VÍNCULO EMPREGATÍCIO")
    cabec = ["PROPONENTE", "FUNÇÃO", "EMPRESA", "ADMISSÃO", "RENDA BRUTA", "RENDA LÍQUIDA", "EXTRAS"]
    starts = [(1, 2), (3, 3), (4, 4), (5, 5), (6, 6), (7, 7), (8, 8)]
    r = _row_cells(ws, r, cabec, fill=_FILL_HEADER, font=_FONT_BOLD, starts=starts)
    r = _row_cells(
        ws, r,
        [d.get("prop1_nome"), d.get("prop1_funcao"), d.get("prop1_empresa"),
         d.get("prop1_admissao"), d.get("prop1_renda_bruta"),
         d.get("prop1_renda_liquida"), d.get("prop1_extras")],
        align=_AL_LEFT, starts=starts,
    )
    r = _row_cells(
        ws, r,
        [d.get("prop2_nome"), d.get("prop2_funcao"), d.get("prop2_empresa"),
         d.get("prop2_admissao"), d.get("prop2_renda_bruta"),
         d.get("prop2_renda_liquida"), d.get("prop2_extras")],
        align=_AL_LEFT, starts=starts,
    )
    r += 1

    # --- COMPROMISSOS FINANCEIROS E DESPESAS ---
    r = _section(ws, r, "COMPROMISSOS FINANCEIROS E DESPESAS")
    r = _kv(ws, r, "EMPRÉSTIMOS, FINANCIAMENTOS, CONSÓRCIOS, OUTROS:", d.get("emprestimos"), label_span=3)
    r = _kv(ws, r, "MORADIA:", d.get("moradia_tipo"), label_span=3)
    r = _kv(ws, r, "TRANSPORTES:", d.get("transportes"), label_span=3)
    r += 1

    # --- RELACIONAMENTO CAIXA ---
    r = _section(ws, r, "RELACIONAMENTO CAIXA")
    r = _kv(ws, r, "CONTA:", d.get("conta"), label_span=2, value_span=2)
    r = _kv(ws, r, "CONTA SALÁRIO:", d.get("conta_salario"), label_span=2, value_span=2)
    r = _kv(ws, r, "OPEN FINANCE:", d.get("open_finance"), label_span=2, value_span=2)
    r = _kv(ws, r, "OPT IN:", d.get("opt_in"), label_span=2, value_span=2)
    r = _kv(ws, r, "BIOMETRIA:", d.get("biometria"), label_span=2, value_span=2)
    r = _kv(ws, r, "CARTÃO CRÉDITO:", d.get("cartao_credito"), label_span=2, value_span=2)
    r = _kv(ws, r, "CROT:", d.get("crot"), label_span=2, value_span=2)
    r += 1

    # --- DADOS DO FINANCIAMENTO ---
    r = _section(ws, r, "DADOS DO FINANCIAMENTO IMÓVEL/CONSTRUÇÃO (PMCMV)")
    r = _kv(ws, r, "VALOR TOTAL:", d.get("valor_total"), label_span=2)
    r = _kv(ws, r, "SUBSÍDIO:", d.get("subsidio"), label_span=2)
    r = _kv(ws, r, "ENTRADA:", d.get("entrada"), label_span=2)
    r = _kv(ws, r, "NEGOCIAÇÃO:", d.get("negociacao"), label_span=2)
    r = _kv(ws, r, "FINANCIADO:", d.get("financiado"), label_span=2)
    r = _kv(ws, r, "PARCELA:", d.get("parcela"), label_span=2)
    r = _kv(ws, r, "PRAZO:", d.get("prazo"), label_span=2)
    r = _kv(ws, r, "AMORTIZAÇÃO:", d.get("amortizacao"), label_span=2)
    r = _kv(ws, r, "UTILIZAR FGTS:", d.get("utilizar_fgts"), label_span=2)
    r += 1

    # --- MORADIA / IMÓVEL ---
    r = _section(ws, r, "MORADIA / IMÓVEL")
    r = _kv(ws, r, "ENDEREÇO DO IMÓVEL:", d.get("endereco_imovel"), label_span=2)
    r = _kv(ws, r, "PROPRIETÁRIOS:", d.get("proprietarios"), label_span=2)
    r = _kv(ws, r, "CONSTRUTORA:", d.get("construtora"), label_span=2)
    r = _kv(ws, r, "PROPRIETÁRIOS (CONSTRUTORA):", d.get("proprietarios_construtora"), label_span=2)
    r += 1

    # --- TAXAS E SEGUROS ---
    r = _section(ws, r, "INFORMAÇÕES SOBRE TAXAS E SEGUROS")
    r = _kv(ws, r, "TAXA À VISTA DO CONTRATO:", d.get("taxa_vista_contrato"), label_span=2)
    r = _kv(ws, r, "SEGURIDADE:", d.get("seguridade"), label_span=2)
    r = _kv(ws, r, "MEMBROS DO COMITÊ:", "", label_span=2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ===========================================================================
# Relatório Contábil consolidado (vendas + comissões + documentos)
# ===========================================================================
def gerar_xlsx_relatorio_contabil(rows: list, filtros: dict = None) -> bytes:
    """Gera XLSX consolidado de oportunidades para a contabilidade.

    Uma linha por oportunidade (colunas de fechamento/comissão) + linha de
    TOTAIS. ``filtros`` é exibido no cabeçalho. Genérico p/ qualquer pipeline.
    """
    from app.services.pdf_service import _format_cpf
    filtros = filtros or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório Contábil"

    cols = [
        ("Cliente", 26), ("CPF", 16), ("Empreendimento", 22), ("Unidade", 12),
        ("Nº Contrato", 14), ("Mês Fech.", 11), ("Data Fech.", 12),
        ("Valor Imóvel", 15), ("Valor Entrada", 14), ("Valor Financ.", 15),
        ("Comissão Total", 15), ("1º Receb.", 13), ("Restante", 13),
        ("Forma Receb.", 14), ("Previsão Receb.", 14), ("Status Comissão", 16),
        ("Responsável", 24), ("Funil", 18), ("Etapa", 20), ("Status", 10),
    ]
    n = len(cols)
    money_cols = {8, 9, 10, 11, 12, 13}

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    t = ws.cell(row=1, column=1, value="RELATÓRIO CONTÁBIL — VENDAS E DOCUMENTOS")
    t.font = Font(name="Calibri", size=13, bold=True)
    t.alignment = _AL_CENTER

    # Filtros
    fparts = []
    if filtros.get("mes"):
        fparts.append("Mês: %s" % filtros["mes"])
    if filtros.get("ano"):
        fparts.append("Ano: %s" % filtros["ano"])
    if filtros.get("empreendimento"):
        fparts.append("Empreendimento: %s" % filtros["empreendimento"])
    if filtros.get("comissao_status"):
        fparts.append("Status comissão: %s" % filtros["comissao_status"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    fc = ws.cell(row=2, column=1, value="Filtros — " + (" · ".join(fparts) if fparts else "todos"))
    fc.font = Font(name="Calibri", size=9, italic=True)
    fc.alignment = _AL_LEFT

    # Cabeçalho
    hrow = 4
    for i, (label, w) in enumerate(cols, start=1):
        c = ws.cell(row=hrow, column=i, value=label)
        c.fill = _FILL_HEADER
        c.font = _FONT_BOLD
        c.alignment = _AL_CENTER
        c.border = _BORDER
        ws.column_dimensions[get_column_letter(i)].width = w

    def _money(v):
        try:
            return float(v) if v is not None else None
        except (TypeError, ValueError):
            return None

    def _date(v):
        try:
            return v.strftime("%d/%m/%Y") if v else ""
        except AttributeError:
            return str(v) if v else ""

    tots = {c: 0.0 for c in money_cols}
    r = hrow + 1
    for d in rows:
        vals = [
            d.get("cliente_nome") or "", _format_cpf(d.get("cliente_cpf")),
            d.get("empreendimento_nome") or d.get("imovel_nome") or "", d.get("unidade") or "",
            d.get("numero_contrato") or "", d.get("contabil_mes_fechamento") or "",
            _date(d.get("contabil_data_fechamento") or d.get("data_fechamento")),
            _money(d.get("valor_imovel")), _money(d.get("valor_entrada")), _money(d.get("valor_total_financiamento")),
            _money(d.get("comissao_total")), _money(d.get("comissao_recebimento_1")), _money(d.get("comissao_restante")),
            d.get("comissao_forma_recebimento") or "", _date(d.get("comissao_previsao_recebimento")),
            d.get("comissao_status") or "", d.get("proprietario_email") or "",
            d.get("pipeline_nome") or "", d.get("stage_nome") or "", d.get("status") or "",
        ]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = _BORDER
            c.font = _FONT_BASE
            if i in money_cols:
                c.number_format = 'R$ #,##0.00'
                c.alignment = _AL_CENTER
                tots[i] += _money(v) or 0
            else:
                c.alignment = _AL_LEFT
        r += 1

    # Totais
    c0 = ws.cell(row=r, column=1, value="TOTAIS (%d)" % len(rows))
    c0.font = _FONT_BOLD
    for i in range(1, n + 1):
        c = ws.cell(row=r, column=i)
        c.border = _BORDER
        c.fill = _FILL_HEADER
        c.font = _FONT_BOLD
        if i in money_cols:
            c.value = tots[i]
            c.number_format = 'R$ #,##0.00'
            c.alignment = _AL_CENTER

    ws.freeze_panes = "A5"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
